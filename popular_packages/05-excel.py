import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

sheet = wb.active

wb.create_sheet("Sheet3")

sheet3 = wb["Sheet3"]
sheet3.title = "New Title"

# print(
#     sheet.max_row,
#     sheet.max_column
# )

cell = sheet["A1"]
cell.value = "Complete name"
# print(cell.value)

cell2 = sheet.cell(row=2, column=1)
# print(
#     cell2.value,
#     cell2.row,
#     cell2.column,
#     cell.coordinate
# )

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=column)
        # print(row, column, cell.value)

column = sheet["A"]
row = sheet["1"]
# print(row)
# print(column)

sheet.append([1, 2, 3])
print(sheet.rows)
sheet.delete_rows(1, 1)

wb.save("new_planilla.xlsx")
